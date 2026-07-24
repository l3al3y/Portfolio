"""
Modul Penjejak Excel (Excel Job Application Tracker)
=====================================================
Modul ini bertanggungjawab mengeksport dan menyelaraskan (sync) rekod
pangkalan data SQLite serta log emel ke dalam fail Microsoft Excel (.xlsx).

Ciri-Ciri Utama:
  1. Multi-Worksheet Dashboard - Menjana 3 helaian utama:
     - 'Applications Monitor' (Senarai penuh status permohonan kerja)
     - 'Summary Dashboard' (Statistik KPI ringkasan & pilar teknikal)
     - 'Email Activity Log' (Jejak audit emel keluar & masuk)
  2. Styling Profesional - Warna tajuk utama (Navy Blue #1F4E78), format nombor
     peratusan, dan penyerlahan warna status (Hijau untuk Applied/Interview).
  3. Auto Column Widths - Melaraskan lebar kolum secara automatik supaya mesra dibaca.
"""

from __future__ import annotations
import sqlite3
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("job_agent.excel")


class ExcelTracker:
    def __init__(self, output_path: str | Path = "JobTracker.xlsx") -> None:
        self.output_path = Path(output_path)

    def sync_from_db(self, db_path: str | Path = "job_agent.db", email_logs: Optional[List[Any]] = None) -> Path:
        """
        Membaca pangkalan data SQLite dan mengeksport ke fail Excel (.xlsx).
        """
        db_path = Path(db_path)
        logger.info("[EXCEL TRACKER] Mengeksport data dari SQLite (%s) ke Excel (%s)...", db_path, self.output_path)

        wb = openpyxl.Workbook()
        # Buang sheet lalai
        wb.remove(wb.active)

        # 1. Helaian Monitor Utama
        ws_main = wb.create_sheet(title="Applications Monitor")
        self._build_main_sheet(ws_main, db_path)

        # 2. Helaian Dashboard Ringkasan
        ws_dash = wb.create_sheet(title="Summary Dashboard")
        self._build_dashboard_sheet(ws_dash, db_path)

        # 3. Helaian Log Emel
        ws_email = wb.create_sheet(title="Email Activity Log")
        self._build_email_sheet(ws_email, email_logs or [])

        # Simpan fail Excel
        wb.save(self.output_path)
        logger.info("[EXCEL TRACKER] BERJAYA MENJANA FAIL EXCEL: %s", self.output_path.resolve())
        return self.output_path

    # ------------------------------------------------------------------
    # HELAIAN 1: APPLICATIONS MONITOR
    # ------------------------------------------------------------------

    def _build_main_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, db_path: Path) -> None:
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "Job ID", "Company", "Job Title", "ATS Match Score (%)",
            "Dominant Domain", "Application Status", "Applied / Processed Date",
            "Cover Letter / Snippet", "Job URL"
        ]

        # Style Tajuk (Header)
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)
        ws.row_dimensions[1].height = 28

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Style Warna Status
        status_styles = {
            "INTERVIEW_INVITE": {
                "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                "font": Font(name="Calibri", size=11, bold=True, color="375623")
            },
            "APPLIED": {
                "fill": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                "font": Font(name="Calibri", size=11, bold=True, color="1F4E78")
            },
            "SKIPPED_LOW_MATCH": {
                "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                "font": Font(name="Calibri", size=11, color="7F6000")
            },
            "SKIPPED_DUPLICATE": {
                "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
                "font": Font(name="Calibri", size=11, color="595959")
            },
            "FAILED": {
                "fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
                "font": Font(name="Calibri", size=11, bold=True, color="C65911")
            }
        }

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        if db_path.is_file():
            with sqlite3.connect(db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute(
                    """
                    SELECT job_id, company, title, match_score, status,
                           cover_letter, created_at, url
                    FROM applied_jobs
                    ORDER BY id DESC
                    """
                )
                rows = cursor.fetchall()

                for row_idx, r in enumerate(rows, start=2):
                    match_score_val = float(r["match_score"]) if r["match_score"] is not None else 0.0
                    status_val = r["status"]

                    domain_val = self._guess_domain_from_title(r["title"])

                    snippet = (r["cover_letter"] or "").strip()
                    if snippet and len(snippet) > 80:
                        snippet = snippet[:77] + "..."

                    ws.append([
                        r["job_id"],
                        r["company"],
                        r["title"],
                        round(match_score_val, 1),
                        domain_val,
                        status_val,
                        r["created_at"],
                        snippet,
                        r["url"]
                    ])

                    ws.row_dimensions[row_idx].height = 22

                    # Format sel
                    score_cell = ws.cell(row=row_idx, column=4)
                    score_cell.alignment = Alignment(horizontal="right")
                    score_cell.number_format = '0.0"%"'

                    status_cell = ws.cell(row=row_idx, column=6)
                    status_cell.alignment = Alignment(horizontal="center")
                    if status_val in status_styles:
                        status_cell.fill = status_styles[status_val]["fill"]
                        status_cell.font = status_styles[status_val]["font"]

                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=c).border = thin_border

        self._auto_fit_columns(ws)

    # ------------------------------------------------------------------
    # HELAIAN 2: SUMMARY DASHBOARD
    # ------------------------------------------------------------------

    def _build_dashboard_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, db_path: Path) -> None:
        ws.views.sheetView[0].showGridLines = True

        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        ws.cell(row=1, column=1, value="JOB APPLICATION AGENT MONITORING DASHBOARD").font = title_font

        # Ambil statistik
        stats: Dict[str, Any] = {}
        total_count = 0
        avg_overall = 0.0

        if db_path.is_file():
            with sqlite3.connect(db_path) as conn:
                cur = conn.execute(
                    "SELECT status, COUNT(*) as cnt, AVG(match_score) as avg_score FROM applied_jobs GROUP BY status"
                )
                for r in cur.fetchall():
                    stats[r[0]] = {"count": r[1], "avg_score": r[2] or 0.0}
                    total_count += r[1]

                cur2 = conn.execute("SELECT AVG(match_score) FROM applied_jobs WHERE status='APPLIED'")
                res = cur2.fetchone()
                avg_overall = res[0] or 0.0

        # Jadual KPI Utama
        kpi_headers = ["Metric / Status", "Total Jobs", "Average Match Score (%)"]
        ws.cell(row=3, column=1, value="OVERALL METRICS").font = Font(size=12, bold=True, color="1F4E78")

        for c, h in enumerate(kpi_headers, start=1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        kpi_rows = [
            ("Total Jobs Processed", total_count, round(avg_overall, 1)),
            ("Applied Successfully", stats.get("APPLIED", {}).get("count", 0), round(stats.get("APPLIED", {}).get("avg_score", 0.0), 1)),
            ("Interview Invites Received", stats.get("INTERVIEW_INVITE", {}).get("count", 0), 100.0 if stats.get("INTERVIEW_INVITE") else 0.0),
            ("Skipped (Low Match Score)", stats.get("SKIPPED_LOW_MATCH", {}).get("count", 0), round(stats.get("SKIPPED_LOW_MATCH", {}).get("avg_score", 0.0), 1)),
            ("Skipped (Duplicate Company)", stats.get("SKIPPED_DUPLICATE", {}).get("count", 0), 0.0),
            ("Failed / Errors", stats.get("FAILED", {}).get("count", 0), 0.0),
        ]

        for r_idx, (label, count, avg) in enumerate(kpi_rows, start=5):
            ws.cell(row=r_idx, column=1, value=label)
            ws.cell(row=r_idx, column=2, value=count).alignment = Alignment(horizontal="right")
            score_c = ws.cell(row=r_idx, column=3, value=avg)
            score_c.alignment = Alignment(horizontal="right")
            score_c.number_format = '0.0"%"'

        # Jadual Pilar Teknikal
        ws.cell(row=13, column=1, value="TECHNICAL PILLAR COVERAGE (CANDIDATE MEMORY)").font = Font(size=12, bold=True, color="1F4E78")
        pillar_headers = ["Technical Pillar", "Key Credentials / Certifications", "Target ATS Keywords Matched"]
        for c, h in enumerate(pillar_headers, start=1):
            cell = ws.cell(row=14, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

        pillars = [
            ("Networking & Infrastructure", "CCNA Enterprise, CCNA Wireless", "CCNA, Cisco, VLAN, OSPF, TCP/IP, WAN/LAN"),
            ("Automation & Industrial AI", "Festo Professional AI Certificate", "Python, OpenCV, YOLOv8, Predictive Maintenance"),
            ("IT Technical Support", "Desktop Support & Hardware", "Windows, Hardware Troubleshooting, User Support"),
            ("Security & Incident Response", "Cisco Endpoint Security, Threat Mgmt", "Endpoint Security, Cybersecurity, Incident Response"),
            ("Embedded Systems & IoT", "INOTEK 2025 3rd Place Award", "Arduino, Microcontroller, Load Cell/HX711, C/C++"),
        ]

        for r_idx, (p_name, certs, kws) in enumerate(pillars, start=15):
            ws.cell(row=r_idx, column=1, value=p_name)
            ws.cell(row=r_idx, column=2, value=certs)
            ws.cell(row=r_idx, column=3, value=kws)

        self._auto_fit_columns(ws)

    # ------------------------------------------------------------------
    # HELAIAN 3: EMAIL ACTIVITY LOG
    # ------------------------------------------------------------------

    def _build_email_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, email_logs: List[Any]) -> None:
        ws.views.sheetView[0].showGridLines = True

        headers = ["Direction", "Sender", "Recipient", "Subject", "Category", "Matched Company", "Timestamp"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        ws.row_dimensions[1].height = 26

        for r_idx, item in enumerate(email_logs, start=2):
            direction = getattr(item, "direction", "OUTGOING")
            sender = getattr(item, "sender", "")
            recipient = getattr(item, "recipient", "")
            subject = getattr(item, "subject", "")
            category = getattr(item, "category", "")
            company = getattr(item, "company_matched", "") or ""
            ts = getattr(item, "timestamp", datetime.now()).strftime("%Y-%m-%d %H:%M")

            ws.append([direction, sender, recipient, subject, category, company, ts])

            # Warna mengikut kategori emel
            cat_cell = ws.cell(row=r_idx, column=5)
            if category == "INTERVIEW_INVITE":
                cat_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                cat_cell.font = Font(bold=True, color="375623")
            elif category == "APPLICATION_SENT":
                cat_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        self._auto_fit_columns(ws)

    # ------------------------------------------------------------------
    # UTILITIES
    # ------------------------------------------------------------------

    @staticmethod
    def _guess_domain_from_title(title: str) -> str:
        t = title.lower()
        if any(w in t for w in ["network", "ccna", "cisco", "infrastructure", "wan", "lan"]):
            return "Networking"
        if any(w in t for w in ["ai", "vision", "yolo", "automation", "opencv", "robotics"]):
            return "Automation & AI"
        if any(w in t for w in ["support", "helpdesk", "desktop", "it executive"]):
            return "IT Support"
        if any(w in t for w in ["security", "cyber", "threat", "soc"]):
            return "Security"
        if any(w in t for w in ["iot", "embedded", "arduino", "hardware"]):
            return "Embedded/IoT"
        return "General IT/Eng"

    @staticmethod
    def _auto_fit_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if val_str:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
